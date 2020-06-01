from PIL import ImageTk, Image           #ALL LIBRARIES
import tkinter as tk
import datetime, random
import openpyxl as op
import requests
from io import BytesIO
import mysql.connector
import os
from dotenv import load_dotenv


"""pictures"""
img_yes_url = requests.get('https://www.derjogger.de/wp-content/uploads/2016/03/Zitat-zu-Ausdauer-Aussehen-und-Sport.png')
img_yes = Image.open(BytesIO(img_yes_url.content))
img_yes.thumbnail((350, 350), Image.ANTIALIAS)

img_no_url = requests.get('https://i.pinimg.com/236x/9c/ec/fc/9cecfcf67576684441228e5b8054b5fb.jpg')
img_no = Image.open(BytesIO(img_no_url.content))
img_no.thumbnail((350, 350), Image.ANTIALIAS)

img_bg_url = requests.get('https://blog.tutoronline.ru/media/621836/learn-1.jpg')
img_bg= Image.open(BytesIO(img_bg_url.content))


class GermanGame:
    #init Gui fro the app
    def __init__(self, root, pick_phrase_func):
        # self.root = root
        self.main_func = pick_phrase_func
        
        root.geometry('1000x500')
        root.configure(background='cyan')
        root.title('Phrase Game')

        image_background = ImageTk.PhotoImage(img_bg)
        label_bg_img = tk.Label(master=root, image=image_background)
        label_bg_img.image = image_background
        label_bg_img.place(x=0, y=0, relwidth=1, relheight=1)

        button_pick_a_phrase = tk.Button(text='Pick a phrase', bg='gold', font=(15))
        button_pick_a_phrase.place(x=220, y=150)
        button_pick_a_phrase.bind('<Button-1>', self.main_func)

        label_version = tk.Label(text='Version 0.2', font=('arial 10 bold'))
        label_version.place(x=900, y=450)

        label_lets_start = tk.Label(master=root, text='LET\'S START', font=('Helvetica', 60, 'bold'), bg='gold')
        label_lets_start.place(x=30, y=10)

        root.mainloop
        
    #gui after clicking on PICK_A_PHRASE button
    def pick_phrase_gui(self, phrase_text, phrase_id, phrase_frequency, yes, no, yes_score, no_score, delete, add):
        text_phrase = tk.Text(width=40, height=5, bg='light goldenrod')
        text_phrase.place(x=120, y=190)
        text_phrase.insert(tk.END, phrase_text)

        button_yes = tk.Button(text='Yes', font=(15), bg='green', command=lambda: yes(phrase_id, phrase_frequency))
        button_yes.place(x=220, y=290)
        # button_yes.bind('<Button-1>', lambda: yes(phrase_id))

        button_no = tk.Button(text='No', font=(15), bg='red', command=lambda: no(phrase_id, phrase_frequency))
        button_no.place(x=300, y=290)
        #no_button.bind('<Button-Down>', no_button)
        #master=self.root,
        button_delete_phrase = tk.Button(text='delete', font=(10), command=lambda: delete(phrase_id))
        button_delete_phrase.place(x=750, y=450)

        button_add = tk.Button(text='add phrase', font=(10))
        button_add.place(x=580, y=450)

        # add_text = tk.Entry(master=self.root)            #not finished yet
        # add_text.place(x=480, y=450)

        label_score = tk.Label(text='Score', bg='gold')
        label_score.place(x=840, y=30)

        text_yes_score = tk.Text(width=4, heigh=1, bg='green', font=('bold'))
        text_yes_score.place(x=790, y=30)
        text_yes_score.insert(tk.END, str(yes_score))

        text_no_score = tk.Text(width=4, height=1, bg='red', font=('bold'))
        text_no_score.place(x=883, y=30)
        text_no_score.insert(tk.END, str(no_score))

    # def add_gui(self):
    #     self.


# def yes_button():
#     global anti_repeat, yes_score
#     if anti_repeat is False:            #prevents YES button spamming with anti_repeat Boolean
#         anti_repeat = True
#         work_sheet.cell(row=n, column=5).value = work_sheet.cell(row=n, column=5).value * 2
#         yes_score += 1

#         image_yes = ImageTk.PhotoImage(img_yes) #ImageTk needs to be inside the root projection
#         label_yes_img = tk.Label(master=root, image=image_yes)
#         label_yes_img.image = image_yes         #an extra reference in order to work properly
#         label_yes_img.place(x=600, y=130)

#         label_yes_img.after(3000, lambda: label_yes_img.destroy())

#         wb.save(file_name)

# def no_button():
#     global anti_repeat, no_score
#     if anti_repeat == False:
#         anti_repeat = True
#         no_score += 1

#         if  work_sheet.cell(row=n, column=5).value >= 2:
#             work_sheet.cell(row=n, column=5).value = work_sheet.cell(row=n, column=5).value / 2

#         help_text = tk.Text(master=root, width=40, height=3, bg='light goldenrod')
#         help_text.place(x=120, y =360)
#         help_text.insert(tk.END, work_sheet['G' + str(n)].value)

#         help_text.after(3000, lambda: help_text.destroy())

#         image_no = ImageTk.PhotoImage(img_no)
#         label_no_img = tk.Label(master=root, image=image_no)
#         label_no_img.image = image_no
#         label_no_img.place(x=600, y=130)

#         label_no_img.after(3000, lambda: label_no_img.destroy())