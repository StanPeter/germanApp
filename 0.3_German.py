from PIL import ImageTk, Image           #ALL LIBRARIES
import tkinter as tk
import datetime, random
import requests
import openpyxl as op
from io import BytesIO
import mysql.connector
from db import db

#connection for MySQL
mydb = mysql.connector.connect(
    host = db["host"],
    user = db["user"],
    passwd = db["password"]
)
print(mydb)

now = datetime.date.today()

file_name = 'database.xlsx'
wb = op.load_workbook(file_name)            #loading from xlsx document
work_sheet = wb['Work_Sheet']
rows_num = work_sheet.max_row

phrase_generator_list = []                  #used to prevent in repeating the same phrases
phrases_counter = 1                          #tracks how many phrases left, set on 1 because of excel start position on 2
no_button_counter = 0                       #not used yet
yes_button_counter = 0                      #not used yet
anti_repeat = False                         #prevents yes/no button spamming
yes_score = 0
no_score = 0
yes_img_list = []                           #not used yet --> next version
no_img_list = []                            #not used yet

img_yes_url = requests.get('https://www.derjogger.de/wp-content/uploads/2016/03/Zitat-zu-Ausdauer-Aussehen-und-Sport.png')
img_yes = Image.open(BytesIO(img_yes_url.content))
img_yes.thumbnail((350, 350), Image.ANTIALIAS)

img_no_url = requests.get('https://i.pinimg.com/236x/9c/ec/fc/9cecfcf67576684441228e5b8054b5fb.jpg')
img_no = Image.open(BytesIO(img_no_url.content))
img_no.thumbnail((350, 350), Image.ANTIALIAS)

img_bg_url = requests.get('https://blog.tutoronline.ru/media/621836/learn-1.jpg')
img_bg= Image.open(BytesIO(img_bg_url.content))

def pick_a_phrase(none):
    global anti_repeat, phrases_counter
    def yes_button():
        global anti_repeat, yes_score
        if anti_repeat == False:            #prevents YES button spamming with anti_repeat Boolean
            anti_repeat = True
            work_sheet.cell(row=n, column=5).value = work_sheet.cell(row=n, column=5).value * 2
            yes_score += 1

            image_yes = ImageTk.PhotoImage(img_yes) #ImageTk needs to be inside the root projection
            label_yes_img = tk.Label(master=root, image=image_yes)
            label_yes_img.image = image_yes         #an extra reference in order to work properly
            label_yes_img.place(x=600, y=130)

            label_yes_img.after(3000, lambda: label_yes_img.destroy())

            wb.save(file_name)

    def no_button():
        global anti_repeat, no_score
        if anti_repeat == False:
            anti_repeat = True
            no_score += 1

            if  work_sheet.cell(row=n, column=5).value >= 2:
                work_sheet.cell(row=n, column=5).value = work_sheet.cell(row=n, column=5).value / 2

            help_text = tk.Text(master=root, width=40, height=3, bg='light goldenrod')
            help_text.place(x=120, y =360)
            help_text.insert(tk.END, work_sheet['G' + str(n)].value)

            help_text.after(3000, lambda: help_text.destroy())

            image_no = ImageTk.PhotoImage(img_no)
            label_no_img = tk.Label(master=root, image=image_no)
            label_no_img.image = image_no
            label_no_img.place(x=600, y=130)

            label_no_img.after(3000, lambda: label_no_img.destroy())

            wb.save(file_name)

    def delete_phrase():
        print('Deleted')
        work_sheet.cell(row=n, column=1).value = None
        work_sheet.cell(row=n, column=3).value = None
        work_sheet.cell(row=n, column=5).value = None
        work_sheet.cell(row=n, column=7).value = None
        wb.save()

    def add_phrase():
        for i in range(2, rows_num):
            if work_sheet.cell(column=1, row=i).value == None:
                if add_text != None:
                    work_sheet.cell(column=1, row=i).value = str(add_text.get())
                    print('Added')
                    wb.save()
                else:
                    print('continue 1')
                    continue

            else:
                print('continue 2')
                continue


    while True:
        n = random.randint(2, rows_num)         #goes over all the phrases in the database
        anti_repeat = False                     #resets the anti_repeat Boolean
        if phrases_counter >= rows_num:
            print('CODE FINISHED')
            break
        if n not in phrase_generator_list:
            phrase_generator_list.append(n)

            if work_sheet['A' + str(n)].value != None:            #Start of the main function

                if work_sheet.cell(row=n, column=3).value == None:
                    work_sheet.cell(row=n, column=5).value = 1
                    work_sheet.cell(row=n, column=3).value = now  #if the cell doesnt have any date puts today's date
                    phrases_counter += 1
                else:
                    datum_counter_minor = now - datetime.datetime.date(work_sheet.cell(row=n, column=3).value)
                    datum_counter_major = datum_counter_minor.days
                    phrases_counter += 1
                    if datum_counter_major >= work_sheet.cell(row=n, column=5).value:  #checks frequency of practising the phrase
                        work_sheet.cell(row=n, column=3).value = now
                        print('pass' + str(n))
                        pass
                    else:
                        print('continue' + str(n))
                        continue

                text_phrase = tk.Text(master=root, width=40, height=5, bg='light goldenrod')
                text_phrase.place(x=120, y=190)
                text_phrase.insert(tk.END, work_sheet['A' + str(n)].value)

                button_yes = tk.Button(master=root, text='Yes', font=(15), bg='green', command=yes_button)
                button_yes.place(x=220, y=290)
                #yes_button.bind('<Up>', yes_button)

                button_no = tk.Button(master=root, text='No', font=(15), bg='red', command=no_button)
                button_no.place(x=300, y=290)
                #no_button.bind('<Button-Down>', no_button)

                button_delete_phrase = tk.Button(master=root, text='delete phrase', font=(10), command=delete_phrase)
                button_delete_phrase.place(x=750, y=450)

                button_add = tk.Button(master=root, text='add phrase', font=(10), command=add_phrase)
                button_add.place(x=650, y=450)

                add_text = tk.Entry(master=root)            #not finished yet
                add_text.place(x=480, y=450)

                label_score = tk.Label(master=root, text='Score', bg='gold')
                label_score.place(x=840, y=30)

                text_yes_score = tk.Text(master=root, width=4, heigh=1, bg='green', font=('bold'))
                text_yes_score.place(x=790, y=30)
                text_yes_score.insert(tk.END, yes_score)

                text_no_score = tk.Text(master=root, width=4, height=1, bg='red', font=('bold'))
                text_no_score.place(x=883, y=30)
                text_no_score.insert(tk.END, no_score)

                wb.save(file_name)
                break
            else:
                print('skipped {}'.format('A'+ str(n)))
                phrases_counter += 1
                continue
        else:
            #print('r')
            continue

#TKINTER GRAPHICS GUY FOR THE APP
root = tk.Tk()
root.geometry('1000x500')
root.configure(background='cyan')
root.title('Phrase Game')

image_background = ImageTk.PhotoImage(img_bg)
label_bg_img = tk.Label(master=root, image=image_background)
label_bg_img.image = image_background
label_bg_img.place(x=0, y=0, relwidth=1, relheight=1)

button_pick_a_phrase = tk.Button(text='Pick a phrase', bg='gold', font=(15))
button_pick_a_phrase.place(x=220, y=150)
button_pick_a_phrase.bind('<Button-1>', pick_a_phrase)

label_version_02 = tk.Label(text='Version 0.2', font=('arial 10 bold'))
label_version_02.place(x=900, y=450)

label_lets_start = tk.Label(master=root, text='LET\'S START', font=('Helvetica', 60, 'bold'), bg='gold')
label_lets_start.place(x=30, y=10)

root.mainloop()
